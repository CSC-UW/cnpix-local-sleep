"""Supplementary Table S1b: companion *data* workbook for Table S1a.

Supplementary Table S1a reports mixed-model contrasts across the six conditions.
This companion dumps the actual per-(subject, structure, condition) values those
models were fit on: OFF rate, OFF count, total OFFness, and every other
OFF-period property S1a reports. One matrix per property, blocked by the three OFF
partitions (All OFFs / Medium + Large / Small OFFs), plus cortical delta power.

It is a formatted recapitulation of ``inst/extdata/summarized_full48h_*_offs.parquet``
(the exact files that feed the S1a fits) and ``summarized_full48h_bandpower_offs.parquet``
(delta power). The property-to-column mapping per partition is read straight from
``config/summary_tables/manuscript_s1a_homeostasis.yaml`` so this stays in lock-step
with S1a: if S1a changes which summary statistic it uses for a partition, this follows.

Separate deliverable, NOT part of ``manuscript_supplement.*``:
    _output_manuscript/manuscript_s1b_values.xlsx   (one sheet per property)
    _output_manuscript/manuscript_s1b_values.csv    (tidy long form, for reuse)

This workbook also backs manuscript Table 2 (confirmed). Table 2 is a
caption-only placeholder in the manuscript ``.docx``, "OFF counts per subject and
structures across the 6 experimental conditions", cited three times for
anatomical-gradient claims, assembled by hand from the ``off_count`` column here:
29 subject-structure pairs x 6 conditions, blocked by the three OFF partitions
(522 rows). Two consequences worth keeping in view:

  * This script and its inputs are on the reproduction spine for two
    deliverables, not one.
  * Because the property-to-column mapping is read from
    ``manuscript_s1a_homeostasis.yaml`` (see above), changing which summary
    statistic S1a uses for a partition propagates into S1b and therefore into
    Table 2. ``off_count`` itself is the ``count`` column in every partition, so
    it is insulated from statistic changes, but not from changes to which
    partitions or which subject-structure rows S1a includes.

Missing-cell semantics (verified from the data): LFP and condition durations exist
for all 29 x 6 cells, and every (subject, structure, condition) cell absent from an
OFF parquet is a Wake condition, i.e. zero OFF periods were detected there under
that criterion, not "unrecorded". So OFF count, OFF rate and Total OFFness are 0 at
those cells (count/time densities of an empty set), while median-based properties
(duration, span, area, residual MUA, transition asynchrony) are undefined -> blank.
The S1a mixed models were fit only on cells with >= 1 OFF, so those 0-cells appear
here for completeness but were not part of the fits.

Subjects are shown with the anonymized IDs used across the cnpix_local_sleep publication
figures (Subject1..Subject15; see ``subject_aliases``); conditions use the external
names (Early/Late baseline, Early/Late SD, Early/Late recovery). Units are given on
each property tab (its subtitle), not on the Notes sheet.

Usage:
    python manuscript_s1b_values.py                # -> _output_manuscript/manuscript_s1b_values.{xlsx,csv}
    python manuscript_s1b_values.py --xlsx out.xlsx --csv out.csv
"""

from __future__ import annotations

import argparse
import math
from collections import OrderedDict
from pathlib import Path

import pandas as pd
import yaml

from plot_results import join_adjusted_edge_statistics

HERE = Path(__file__).resolve().parent
S1A_CONFIG = HERE / "config" / "summary_tables" / "manuscript_s1a_homeostasis.yaml"
EXTDATA = HERE / "inst" / "extdata"

# The three OFF partitions in the manuscript, in S1a column order (no BLAS).
SETS = OrderedDict([
    ("llas", "All OFFs"),
    ("clas", "Medium + Large"),
    ("llas_exclusive", "Small OFFs"),
])

# Six conditions in experimental-time order. Internal tokens are used only for
# lookups against the parquets; the sheets and CSV show the external names.
CONDITIONS = [
    ("Early.BSL.NREM", "Early baseline"),
    ("Early.REC.NREM.Match", "Late baseline"),
    ("Early.NOD.Wake", "Early SD"),
    ("Late.NOD.Wake", "Late SD"),
    ("Early.REC.NREM", "Early recovery"),
    ("Late.REC.NREM", "Late recovery"),
]
CONDITION_ORDER = [c for c, _ in CONDITIONS]
CONDITION_LABEL = dict(CONDITIONS)  # internal token -> external name

# Properties whose value at a zero-OFF cell is a well-defined 0 (a count, or a
# per-unit-time density of the empty set) rather than an undefined median -> blank.
ZERO_FILL = {"OFF count", "OFF rate", "Total OFFness"}

# Sheet display order: the two the manuscript companion leads with, then S1a order.
DISPLAY_ORDER = [
    "OFF count", "OFF rate", "Total OFFness", "OFF duration", "OFF span",
    "OFF area (per-OFF)", "Residual MUA",
    # Reported for Medium+Large only, so these sheets carry one block where the
    # others carry three (see the S1a config for why the other partitions are not
    # fitted, and why the unadjusted MAD rows are no longer reported at all).
    "ON-transition asynchrony (size-adjusted)",
    "OFF-transition asynchrony (size-adjusted)",
]

# Plain-language unit shown in each property sheet's subtitle (units live on the
# tabs, not the Notes sheet).
UNIT_PHRASE = {
    "OFF count": "Counts (number of OFF periods)",
    "OFF rate": "Hz (OFF periods per second)",
    "Total OFFness": "Arbitrary units (a.u.)",
    "OFF duration": "Seconds (s)",
    "OFF span": "Micrometres (µm)",
    "OFF area (per-OFF)": "Micrometre·seconds (µm·s)",
    "Residual MUA": "Microvolts (µV)",
    # The size-adjusted columns are exported in ms, not seconds; see
    # cnpix_local_sleep's export-adjusted-edge-statistics and report section 5.2.
    "ON-transition asynchrony (size-adjusted)": "Milliseconds (ms)",
    "OFF-transition asynchrony (size-adjusted)": "Milliseconds (ms)",
}
DELTA_UNIT_PHRASE = "Z-scored log delta power"


def subject_aliases(subjects) -> dict:
    """Canonical anonymized IDs used across the cnpix_local_sleep publication figures:
    Subject1..SubjectN by lexicographic sort of the internal CNPIX* IDs (so
    'CNPIX10-Charles' -> 'Subject1', 'CNPIX2-Segundo' -> 'Subject10'). Mirrors
    ``{s: f"Subject{i}" for i, s in enumerate(sorted(offs.subject.unique()), 1)}``
    in notebooks/figures/*."""
    return {s: f"Subject{i}" for i, s in enumerate(sorted(set(subjects)), start=1)}

# Excel sheet-name (<= 31 chars) per property label.
SHEET_NAMES = {
    "OFF count": "OFF count",
    "OFF rate": "OFF rate",
    "Total OFFness": "Total OFFness",
    "OFF duration": "OFF duration",
    "OFF span": "OFF span",
    "OFF area (per-OFF)": "OFF area",
    "Residual MUA": "Residual MUA",
    "ON-transition asynchrony (size-adjusted)": "ON-transition async (adj)",
    "OFF-transition asynchrony (size-adjusted)": "OFF-transition async (adj)",
}


def _round_sig(x, sig: int = 4):
    """Round to ``sig`` significant figures, returning a float (so Excel stores a
    real number). None passes through; integers stay integers."""
    if x is None or (isinstance(x, float) and not math.isfinite(x)):
        return None
    if x == 0:
        return 0
    return round(x, -int(math.floor(math.log10(abs(x)))) + (sig - 1))


# Read the S1a config for the property -> per-partition summary column mapping
def load_property_map(s1a_config: dict):
    """Return prop_cols[label][dataset] = the summarized-parquet column S1a uses for
    that (property, partition), derived from S1a's own ``entries`` +
    ``property_labels`` so the companion tracks S1a exactly. The synthetic 'OFF count'
    property (column ``count``, all partitions) is injected because raw counts are
    requested but are not themselves an S1a row."""
    property_labels = s1a_config["property_labels"]
    prop_cols: "OrderedDict[str, dict]" = OrderedDict()
    for section in s1a_config["sections"]:
        for rv, ds, _cset, _model in section["entries"]:
            if ds not in SETS:  # skip the bandpower (delta-power) entry
                continue
            prop_cols.setdefault(property_labels.get(rv, rv), {})[ds] = rv
    prop_cols["OFF count"] = {ds: "count" for ds in SETS}  # same column, all partitions
    return prop_cols


# -------------------- Assemble matrices --------------------
def load_set_tables():
    """Return {dataset: {(subject, structure, condition): row-Series}} for the three
    OFF partitions, plus the sorted list of 29 (subject, structure) combos and a
    per-combo bandpower (delta) lookup."""
    tables, combos = {}, set()
    for ds in SETS:
        df = pd.read_parquet(EXTDATA / f"summarized_full48h_{ds}_offs.parquet")
        # The size-adjusted edge columns live in a companion parquet, so join them
        # in the same way offp::join_adjusted_edge_statistics does for the fits.
        # Without this the S1a rows for size-adjusted asynchrony would have no values
        # behind them in this workbook.
        df = join_adjusted_edge_statistics(df, EXTDATA, ds)
        tables[ds] = {
            (r.subject, r.structure, r.condition): r
            for r in df.itertuples(index=False)
        }
        combos |= {(r.subject, r.structure) for r in df.itertuples(index=False)}
    combos = sorted(combos)  # (subject, structure)
    bp = pd.read_parquet(EXTDATA / "summarized_full48h_bandpower_offs.parquet")
    delta = {
        (r.subject, r.structure, r.condition): r.mean_zlog_delta
        for r in bp.itertuples(index=False)
    }
    return tables, combos, delta


def matrix_for(label, prop_cols, tables, combos, anon):
    """One property's matrix: {dataset: [(anon subject, structure, [val/condition])]}.
    Lookups use the raw subject; the emitted row carries the anonymized ID."""
    zero_fill = label in ZERO_FILL
    is_count = label == "OFF count"
    out = OrderedDict()
    for ds in SETS:
        col = prop_cols[label].get(ds)
        if col is None:  # this partition does not report this property
            continue
        rows = []
        for subject, structure in combos:
            vals = []
            for cond in CONDITION_ORDER:
                r = tables[ds].get((subject, structure, cond))
                if r is None:
                    vals.append(0 if zero_fill else None)
                else:
                    v = getattr(r, col)
                    v = int(v) if is_count else _round_sig(float(v))
                    vals.append(v)
            rows.append((anon[subject], structure, vals))
        out[ds] = rows
    return out


def delta_matrix(delta, combos, anon):
    """Delta-power matrix (single block; not OFF-partition-specific)."""
    rows = []
    for subject, structure in combos:
        vals = [
            _round_sig(delta.get((subject, structure, c)))
            if delta.get((subject, structure, c)) is not None else None
            for c in CONDITION_ORDER
        ]
        rows.append((anon[subject], structure, vals))
    return rows


# -------------------- Emitters --------------------
def _cond_headers():
    return [plain for _tok, plain in CONDITIONS]


NOTES = [
    "Companion data to Supplementary Table S1a.",
    "",
    "n = 29 unique subject-structure pairs (12 cortical structures, 15 rats).",
]


def _style(ws):
    from openpyxl.styles import Alignment, Font, PatternFill, Side

    return {
        "thin": Side(style="thin", color="999999"),
        "center": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "left": Alignment(horizontal="left", vertical="center", wrap_text=True),
        "fill": PatternFill("solid", fgColor="E8E8E8"),
        "bold": Font(bold=True),
        "bolditalic": Font(bold=True, italic=True),
    }


def _fill_notes(ws):
    from openpyxl.styles import Alignment, Font

    ws.cell(1, 1, "Supplementary Table S1b: companion data to Table S1a").font = Font(
        bold=True, size=13)
    for i, line in enumerate(NOTES):
        c = ws.cell(3 + i, 1, line)
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 118


def _fill_matrix(ws, title, blocks, subtitle):
    """blocks = [(block_label_or_None, rows)]; rows = [(subject, structure, [vals])]."""
    from openpyxl.styles import Border, Font

    st = _style(ws)
    headers = ["Subject", "Structure"] + _cond_headers()
    ncol = len(headers)

    ws.cell(1, 1, title).font = Font(bold=True, size=12)
    if subtitle:
        sc = ws.cell(2, 1, subtitle)
        sc.font = Font(italic=True, size=9)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncol)
        sc.alignment = st["left"]
    hdr_row = 3
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(hdr_row, c, h)
        cell.font = st["bold"]
        cell.alignment = st["center"]
        cell.border = Border(bottom=st["thin"], top=st["thin"])
    r = hdr_row + 1

    single = len(blocks) == 1 and blocks[0][0] is None
    for block_label, rows in blocks:
        if not single:
            bc = ws.cell(r, 1, block_label)
            bc.font = st["bolditalic"]
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
            for c in range(1, ncol + 1):
                ws.cell(r, c).fill = st["fill"]
            r += 1
        for subject, structure, vals in rows:
            ws.cell(r, 1, subject).alignment = st["left"]
            ws.cell(r, 2, structure).alignment = st["left"]
            for j, v in enumerate(vals):
                cell = ws.cell(r, 3 + j, v)
                cell.alignment = st["center"]
            r += 1

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 12
    for i in range(len(CONDITIONS)):
        ws.column_dimensions[chr(ord("C") + i)].width = 15
    ws.freeze_panes = f"C{hdr_row + 1}"


def write_xlsx(prop_cols, tables, combos, delta, anon, path: Path):
    from openpyxl import Workbook

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    _fill_notes(wb.active)
    wb.active.title = "Notes"

    for label in DISPLAY_ORDER:
        if label not in prop_cols:
            continue
        mat = matrix_for(label, prop_cols, tables, combos, anon)
        blocks = [(SETS[ds], mat[ds]) for ds in SETS if ds in mat]
        subtitle = f"{UNIT_PHRASE[label]}, per subject and structure across the six conditions."
        ws = wb.create_sheet(title=SHEET_NAMES[label][:31])
        _fill_matrix(ws, label, blocks, subtitle)

    ws = wb.create_sheet(title="Delta power")
    _fill_matrix(
        ws, "Cortical delta power",
        [(None, delta_matrix(delta, combos, anon))],
        f"{DELTA_UNIT_PHRASE}, per subject and structure across the six conditions.",
    )
    wb.save(path)


def write_csv(prop_cols, tables, combos, delta, anon, path: Path):
    """Tidy long form: one row per (partition, subject, structure, condition), with
    anonymized subject IDs and external condition names (no internal identifiers)."""
    import csv

    col_order = [
        ("OFF count", "off_count"),
        ("OFF rate", "off_rate_hz"),
        ("Total OFFness", "total_offness_au"),
        ("OFF duration", "off_duration_s"),
        ("OFF span", "off_span_um"),
        ("OFF area (per-OFF)", "off_area_um_s"),
        ("Residual MUA", "residual_mua_uv"),
        # Medium+Large only, so these are blank for the other two partitions;
        # the same convention this file uses for any undefined property.
        ("ON-transition asynchrony (size-adjusted)", "on_transition_async_ms"),
        ("OFF-transition asynchrony (size-adjusted)", "off_transition_async_ms"),
    ]
    mats = {label: matrix_for(label, prop_cols, tables, combos, anon)
            for label, _ in col_order}
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["off_partition", "subject", "structure", "condition",
                    *[k for _, k in col_order], "delta_power_zlog"])
        for ds, ds_label in SETS.items():
            for i, (subject, structure) in enumerate(combos):
                for jc, cond in enumerate(CONDITION_ORDER):
                    vals = [
                        mats[label][ds][i][2][jc] if ds in mats[label] else ""
                        for label, _ in col_order
                    ]
                    dv = delta.get((subject, structure, cond))
                    w.writerow([ds_label, anon[subject], structure,
                                CONDITION_LABEL[cond], *vals,
                                _round_sig(dv) if dv is not None else ""])


def main(argv=None):
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", type=Path, default=None)
    parser.add_argument("--csv", type=Path, default=None)
    parser.add_argument("--out-dir", type=Path,
                        default=HERE / "_output_manuscript")
    args = parser.parse_args(argv)

    with open(S1A_CONFIG) as f:
        s1a_config = yaml.safe_load(f)
    prop_cols = load_property_map(s1a_config)
    tables, combos, delta = load_set_tables()
    anon = subject_aliases(sub for sub, _ in combos)

    xlsx = args.xlsx or (args.out_dir / "manuscript_s1b_values.xlsx")
    csv_path = args.csv or (args.out_dir / "manuscript_s1b_values.csv")
    write_xlsx(prop_cols, tables, combos, delta, anon, xlsx)
    write_csv(prop_cols, tables, combos, delta, anon, csv_path)
    print(f"n combos: {len(combos)}  properties: {len([p for p in DISPLAY_ORDER if p in prop_cols])}")
    print(f"XLSX: {xlsx}\nCSV : {csv_path}")


if __name__ == "__main__":
    main()
